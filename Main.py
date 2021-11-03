from openpyxl import load_workbook
from openpyxl.workbook import Workbook

## Ruta del archivo base
ruta_base = "Prueba 1.xlsx"

original = load_workbook(ruta_base)
hoja = original["Hoja1"]

rut_listos = []
rut_totales = []


#primero obtenemos a todos los ruts del archivo

for i in range(hoja.max_row):
    if i == 0:
        pass
    else:
        index = "C" + str(i+1)
        if hoja[index].value not in rut_totales:
            rut_totales.append(hoja[index].value)

diccionario_compras = {}

for i in rut_totales:
    diccionario_compras[i] = []

for i in range(hoja.max_row):
    if i == 0:
        pass
    else:
        index = "C" + str(i+1)
        for j in diccionario_compras:
            if j == hoja[index].value:
                diccionario_compras[j].append(i+1)

# El diccionario tiene como llaves los rut y como contenido una lista con los indices
# que tiene ese rut

letras = "ABCDEFGHIJKLMNO"
letras_2 = "ABCDEFGHIJ"
letras_3 = "NO"

folios_procesados = []
for i in diccionario_compras:
    nuevo_archivo = Workbook()
    hoja_nueva = nuevo_archivo.active
    contador = 2
    for letra in letras:
        hoja_nueva[letra + str("1")] = hoja[letra + str("1")].value

    for j in diccionario_compras[i]:
        if hoja["H"+str(j)].value == 0:
            for k in letras:
                hoja_nueva[k+str(contador)] = hoja[k + str(j)].value
            contador += 1

        else:
            comprobador = False
            for k in range(hoja.max_row):
                if hoja["H" + str(j)].value == hoja["H"+str(k+1)].value and j!=k+1:
                    comprobador = True

            if not comprobador:
                for k in letras:
                    hoja_nueva[k + str(contador)] = hoja[k + str(j)].value

            elif comprobador and [hoja["H" + str(j)].value, hoja["J" + str(j)].value] not in folios_procesados:
                folios_procesados.append([hoja["H" + str(j)].value, hoja["J" + str(j)].value])
                cantidad_total = 0
                monto_total = 0
                for k in range(hoja.max_row):
                    h = k+1
                    if [hoja["H" + str(j)].value, hoja["J" + str(j)].value]  == \
                            [hoja["H" + str(h)].value, hoja["J" + str(h)].value] :
                        numero = str(hoja["K"+str(h)].value)
                        numero = numero.split(",")
                        numero = ".".join(numero)
                        cantidad_total += float(numero)
                        precio = str(hoja["L" + str(h)].value)
                        precio = precio.split(",")
                        precio = ".".join(precio)
                        monto = str(hoja["M" + str(h)].value)
                        monto= monto.split(",")
                        monto = ".".join(monto)
                        monto_total += round(float(monto),0)
                precio_promedio = monto_total/cantidad_total
                for k in letras_2:
                    hoja_nueva[k + str(contador)] = hoja[k + str(j)].value
                hoja_nueva["K" +str(contador)] = str(cantidad_total)
                hoja_nueva["L" + str(contador)] = hoja["L" + str(j)].value
                hoja_nueva["M" + str(contador)] = str(cantidad_total * precio_promedio)

                for k in letras_3:
                    hoja_nueva[k + str(contador)] = hoja[k + str(j)].value
                contador += 1




    nuevo_archivo.save(i+".xlsx")







